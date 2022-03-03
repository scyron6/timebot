import re
import os

from openpyxl.styles import Font, Alignment
from openpyxl import Workbook


def add_title_to_worksheet(ws, title):
    ws.append([title])
    cell_string = 'A' + str(ws.max_row)
    c = ws[cell_string]
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True, size=14)
    ws.merged_cells.ranges.append(cell_string + ':B' + str(ws.max_row))


def add_subtitle_to_worksheet(ws, title):
    ws.append([title])
    cell_string = 'A' + str(ws.max_row)
    c = ws[cell_string]
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True, size=11)
    ws.merged_cells.ranges.append(cell_string + ':B' + str(ws.max_row))


def write_total_summary(wb, clients, roles, start_date, end_date):
    ws = wb.create_sheet('Total Summary')
    ws.column_dimensions['A'].width = 40
    summary_title = 'Summary ' + str(start_date) + ' - ' + str(end_date)
    add_title_to_worksheet(ws, summary_title)
    add_title_to_worksheet(ws, 'Total Time By Client')
    for client in clients:
        ws.append([client, clients[client].total_minutes])
    ws.append([])

    add_title_to_worksheet(ws, 'Total Time By Role')
    for role in roles:
        ws.append([role, roles[role]])
    ws.append([])


def write_client_summary(wb, clients):
    ws = wb.create_sheet('Each Client Summary')
    ws.column_dimensions['A'].width = 40
    for client in clients:
        add_subtitle_to_worksheet(ws, client)
        for work in clients[client].minutes_per_work:
            ws.append([work, clients[client].minutes_per_work[work]])
        ws.append([])


def write_workbook(clients, roles, employees, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    write_total_summary(wb, clients, roles, start_date, end_date)
    write_client_summary(wb, clients)

    for employee in employees:
        sheet_name = employee + ' Summary'
        ws = wb.create_sheet(sheet_name)
        ws.column_dimensions['A'].width = 40

        # Print Time by Client
        add_title_to_worksheet(ws, 'Time By Client')
        for client in employees[employee].minutes_per_client:
            ws.append([client, employees[employee].minutes_per_client[client]])
        ws.append([])

        # Print Time By Task
        add_title_to_worksheet(ws, 'Time By Task')
        for task in employees[employee].minutes_per_task:
            ws.append([task, employees[employee].minutes_per_task[task]])
        ws.append([])

        # Print Time By Role
        add_title_to_worksheet(ws, 'Time By Role')
        for role in employees[employee].minutes_per_role:
            ws.append([role, employees[employee].minutes_per_role[role]])
        ws.append([])

        # Print Work for Each Client
        add_title_to_worksheet(ws, 'Work per Client')

        # Print each client
        clients = employees[employee].work_per_client
        for client in clients:
            add_subtitle_to_worksheet(ws, client)
            for work in clients[client]:
                ws.append([work, clients[client][work]])
            ws.append([])

    return(wb)
